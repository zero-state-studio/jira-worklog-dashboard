"""
Core billing logic - rate resolution, worklog aggregation, preview computation, and Excel export.
"""
import io
import json
from datetime import date
from collections import defaultdict
from typing import Optional

from .models import Worklog, BillingPreviewLineItem, BillingPreviewResponse
from .cache import get_storage


def resolve_rate(
    worklog: Worklog,
    billing_project: Optional[dict],
    client: dict,
    rates: list[dict],
    classification: Optional[dict] = None
) -> float:
    """
    Resolve the hourly rate for a worklog using the priority cascade:
    1. Classification override rate (if set)
    2. Specific rate (matching user_email + issue_type + date range)
    3. User-specific rate (matching user_email only)
    4. Issue-type-specific rate (matching issue_type only)
    5. Billing project default rate
    6. Client default rate
    7. Fallback to 0
    """
    # 1. Classification override
    if classification and classification.get("override_hourly_rate") is not None:
        return classification["override_hourly_rate"]

    worklog_date = worklog.started.strftime("%Y-%m-%d") if worklog.started else None
    # Extract issue type from parent_type or fallback
    issue_type = worklog.parent_type

    best_rate = None
    best_specificity = -1

    for rate in rates:
        # Check date validity
        if rate.get("valid_from") and worklog_date and worklog_date < rate["valid_from"]:
            continue
        if rate.get("valid_to") and worklog_date and worklog_date > rate["valid_to"]:
            continue

        specificity = 0
        matches = True

        # Check user match
        if rate.get("user_email"):
            if rate["user_email"].lower() == worklog.author_email.lower():
                specificity += 2
            else:
                matches = False

        # Check issue type match
        if rate.get("issue_type"):
            if issue_type and rate["issue_type"].lower() == issue_type.lower():
                specificity += 1
            else:
                matches = False

        if matches and specificity > best_specificity:
            best_specificity = specificity
            best_rate = rate["hourly_rate"]

    # 2/3/4. Use best matching rate if found
    if best_rate is not None:
        return best_rate

    # 5. Billing project default
    if billing_project and billing_project.get("default_hourly_rate") is not None:
        return billing_project["default_hourly_rate"]

    # 6. Client default
    if client.get("default_hourly_rate") is not None:
        return client["default_hourly_rate"]

    # 7. Fallback
    return 0


async def compute_billing_preview(
    client_id: int,
    period_start: date,
    period_end: date,
    group_by: str = "project",
    billing_project_id: Optional[int] = None
) -> BillingPreviewResponse:
    """
    Compute a live billing preview for a client in a period.
    Groups worklogs and applies rate resolution to produce line items.
    """
    storage = get_storage()

    # Get client
    client = await storage.get_billing_client(client_id)
    if not client:
        raise ValueError(f"Client {client_id} not found")

    # Get billing projects for client (or single project)
    if billing_project_id:
        project = await storage.get_billing_project(billing_project_id)
        if not project or project["client_id"] != client_id:
            raise ValueError(f"Billing project {billing_project_id} not found for client {client_id}")
        projects = [project]
    else:
        projects = await storage.get_billing_projects_by_client(client_id)

    if not projects:
        return BillingPreviewResponse(
            client_id=client_id,
            client_name=client["name"],
            billing_project_id=billing_project_id,
            billing_project_name=projects[0]["name"] if billing_project_id and projects else None,
            period_start=period_start,
            period_end=period_end,
            currency=client["billing_currency"],
            group_by=group_by,
            line_items=[],
            subtotal_amount=0,
            billable_hours=0,
            non_billable_hours=0
        )

    # Build mapping: (jira_instance, project_key) -> billing_project
    mapping_to_project = {}
    for proj in projects:
        for m in proj.get("mappings", []):
            mapping_to_project[(m["jira_instance"], m["jira_project_key"])] = proj

    # Get rates for all projects
    project_rates = {}
    for proj in projects:
        project_rates[proj["id"]] = await storage.get_billing_rates(proj["id"])

    # Get all worklogs in the period for mapped JIRA instances
    from .config import get_users_from_db
    users = await get_users_from_db()
    all_emails = [u["email"] for u in users]

    all_worklogs = await storage.get_worklogs_in_range(
        period_start, period_end, user_emails=all_emails
    )

    # Filter worklogs to those matching our billing project mappings
    matched_worklogs = []
    worklog_project_map = {}  # worklog_id -> billing_project

    for wl in all_worklogs:
        project_key = wl.issue_key.split("-")[0] if "-" in wl.issue_key else wl.issue_key
        key = (wl.jira_instance, project_key)
        if key in mapping_to_project:
            matched_worklogs.append(wl)
            worklog_project_map[wl.id] = mapping_to_project[key]

    # Get classifications for matched worklogs
    worklog_ids = [wl.id for wl in matched_worklogs]
    classifications = await storage.get_worklog_classifications(worklog_ids)

    # Aggregate and compute
    billable_hours = 0
    non_billable_hours = 0

    # Group worklogs based on group_by parameter
    groups = defaultdict(lambda: {"hours": 0, "amount": 0, "description": "", "metadata": {}})

    for wl in matched_worklogs:
        hours = wl.time_spent_seconds / 3600
        proj = worklog_project_map[wl.id]
        classification = classifications.get(wl.id)
        rates = project_rates.get(proj["id"], [])

        # Check if billable (default True)
        is_billable = True
        if classification:
            is_billable = classification["is_billable"]

        if not is_billable:
            non_billable_hours += hours
            continue

        billable_hours += hours
        rate = resolve_rate(wl, proj, client, rates, classification)
        amount = round(hours * rate, 2)

        # Determine group key
        if group_by == "user":
            group_key = wl.author_email.lower()
            # Build email->name lookup
            user_name = wl.author_display_name or wl.author_email
            for u in users:
                if u["email"].lower() == wl.author_email.lower():
                    user_name = f"{u['first_name']} {u['last_name']}"
                    break
            groups[group_key]["description"] = user_name
            groups[group_key]["metadata"] = {"email": wl.author_email}
        elif group_by == "issue":
            group_key = wl.issue_key
            groups[group_key]["description"] = f"{wl.issue_key} - {wl.issue_summary}"
            groups[group_key]["metadata"] = {"issue_key": wl.issue_key, "jira_instance": wl.jira_instance}
        else:  # "project" (default)
            group_key = f"{proj['id']}"
            groups[group_key]["description"] = proj["name"]
            groups[group_key]["metadata"] = {"billing_project_id": proj["id"]}

        groups[group_key]["hours"] += hours
        groups[group_key]["amount"] += amount

    # Build line items
    line_items = []
    for group_key, data in sorted(groups.items(), key=lambda x: x[1]["amount"], reverse=True):
        avg_rate = round(data["amount"] / data["hours"], 2) if data["hours"] > 0 else 0
        line_items.append(BillingPreviewLineItem(
            description=data["description"],
            quantity_hours=round(data["hours"], 2),
            hourly_rate=avg_rate,
            amount=round(data["amount"], 2),
            group_key=group_key,
            metadata=data["metadata"]
        ))

    subtotal = round(sum(li.amount for li in line_items), 2)

    return BillingPreviewResponse(
        client_id=client_id,
        client_name=client["name"],
        billing_project_id=billing_project_id,
        billing_project_name=projects[0]["name"] if billing_project_id and projects else None,
        period_start=period_start,
        period_end=period_end,
        currency=client["billing_currency"],
        group_by=group_by,
        line_items=line_items,
        subtotal_amount=subtotal,
        billable_hours=round(billable_hours, 2),
        non_billable_hours=round(non_billable_hours, 2)
    )


async def generate_invoice_excel(invoice: dict) -> io.BytesIO:
    """
    Generate an Excel file for an invoice.
    Sheet 1: Invoice summary with line items.
    Sheet 2: Detailed worklog data.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()

    # ===== Sheet 1: Fattura =====
    ws = wb.active
    ws.title = "Fattura"

    # Styles
    header_font = Font(bold=True, size=14)
    subheader_font = Font(bold=True, size=11)
    table_header_font = Font(bold=True, size=10, color="FFFFFF")
    table_header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    currency_format = '#,##0.00 €'
    hours_format = '#,##0.00'
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Header
    ws.merge_cells('A1:F1')
    ws['A1'] = f"Fattura #{invoice['id']}"
    ws['A1'].font = header_font

    # Invoice info
    row = 3
    info_pairs = [
        ("Cliente:", invoice.get("client_name", "")),
        ("Progetto:", invoice.get("billing_project_name", "Tutti i progetti")),
        ("Periodo:", f"{invoice['period_start']} - {invoice['period_end']}"),
        ("Stato:", invoice["status"]),
        ("Valuta:", invoice["currency"]),
    ]
    for label, value in info_pairs:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = value
        row += 1

    row += 1

    # Line items table
    ws[f'A{row}'] = "Dettaglio Voci"
    ws[f'A{row}'].font = subheader_font
    row += 1

    # Table headers
    headers = ["#", "Descrizione", "Ore", "Tariffa Oraria", "Importo"]
    col_widths = [6, 50, 12, 16, 16]
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = table_header_font
        cell.fill = table_header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
        ws.column_dimensions[chr(64 + col_idx)].width = width
    row += 1

    # Line items
    line_items = invoice.get("line_items", [])
    for idx, li in enumerate(line_items, 1):
        ws.cell(row=row, column=1, value=idx).border = thin_border
        ws.cell(row=row, column=2, value=li["description"]).border = thin_border
        cell_hours = ws.cell(row=row, column=3, value=li["quantity_hours"])
        cell_hours.number_format = hours_format
        cell_hours.border = thin_border
        cell_rate = ws.cell(row=row, column=4, value=li["hourly_rate"])
        cell_rate.number_format = currency_format
        cell_rate.border = thin_border
        cell_amount = ws.cell(row=row, column=5, value=li["amount"])
        cell_amount.number_format = currency_format
        cell_amount.border = thin_border
        row += 1

    # Totals
    row += 1
    ws.cell(row=row, column=4, value="Subtotale:").font = Font(bold=True)
    cell = ws.cell(row=row, column=5, value=invoice["subtotal_amount"])
    cell.number_format = currency_format
    cell.font = Font(bold=True)
    row += 1

    if invoice.get("taxes_amount", 0) > 0:
        ws.cell(row=row, column=4, value="IVA:").font = Font(bold=True)
        cell = ws.cell(row=row, column=5, value=invoice["taxes_amount"])
        cell.number_format = currency_format
        row += 1

    ws.cell(row=row, column=4, value="TOTALE:").font = Font(bold=True, size=12)
    cell = ws.cell(row=row, column=5, value=invoice["total_amount"])
    cell.number_format = currency_format
    cell.font = Font(bold=True, size=12)

    # Notes
    if invoice.get("notes"):
        row += 2
        ws[f'A{row}'] = "Note:"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        ws[f'A{row}'] = invoice["notes"]

    # ===== Sheet 2: Dettaglio Worklogs =====
    ws2 = wb.create_sheet("Dettaglio Voci")

    # Get detailed worklog data from line items metadata
    ws2_headers = ["#", "Descrizione", "Ore", "Tariffa Oraria", "Importo", "Metadati"]
    ws2_widths = [6, 50, 12, 16, 16, 40]
    for col_idx, (header, width) in enumerate(zip(ws2_headers, ws2_widths), 1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.font = table_header_font
        cell.fill = table_header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
        ws2.column_dimensions[chr(64 + col_idx)].width = width

    for idx, li in enumerate(line_items, 1):
        r = idx + 1
        ws2.cell(row=r, column=1, value=idx).border = thin_border
        ws2.cell(row=r, column=2, value=li["description"]).border = thin_border
        cell_h = ws2.cell(row=r, column=3, value=li["quantity_hours"])
        cell_h.number_format = hours_format
        cell_h.border = thin_border
        cell_r = ws2.cell(row=r, column=4, value=li["hourly_rate"])
        cell_r.number_format = currency_format
        cell_r.border = thin_border
        cell_a = ws2.cell(row=r, column=5, value=li["amount"])
        cell_a.number_format = currency_format
        cell_a.border = thin_border
        metadata_str = ""
        if li.get("metadata_json"):
            try:
                metadata_str = json.dumps(json.loads(li["metadata_json"]), indent=2)
            except (json.JSONDecodeError, TypeError):
                metadata_str = li["metadata_json"] or ""
        ws2.cell(row=r, column=6, value=metadata_str).border = thin_border

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
